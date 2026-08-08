"""Bytes to rows, without trusting any of them. §17.1, §21.

An uploaded spreadsheet is hostile input from a file format designed to carry executable
content. Three separate things are refused here, before openpyxl is asked to read anything:

* **Macros.** A `.xlsm` carries `vbaProject.bin`. openpyxl never executes it — it is a ZIP and
  XML reader with no interpreter in it — but a workbook that contains one has no business in a
  frequency plan, and refusing it means nobody has to reason about what a later library version
  might do with it.
* **Zip bombs.** The central directory declares how large each member expands to. A file that
  claims to expand past :data:`MAX_UNCOMPRESSED_BYTES` is refused without being expanded, which
  is the only point at which refusing it is cheap.
* **External links.** ``keep_links=False`` drops references to other workbooks, so nothing in
  the file can point at a path on the server.

And one thing is refused per *cell*, which is the heart of §17.1: the workbook is opened with
``data_only=False``, so a formula cell yields **the formula** and never Excel's cached answer.
The alternative — ``data_only=True`` — hands over a number some other program computed, at some
other time, possibly against different inputs. That number is exactly what this platform exists
not to depend on.
"""

from __future__ import annotations

import dataclasses
import hashlib
import io
import zipfile
from typing import Any

from openpyxl import load_workbook

from imports_exports.export.safety import restore

#: Ten megabytes of upload. An export of 10⁵ Satnet Paths is well under one, and a
#: frequency-plan workbook that is larger than this is something else.
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

#: What the archive is allowed to claim it expands to. XML compresses roughly tenfold, so this
#: is generous for a real workbook and small enough that a decompression bomb is refused before
#: any of it is read.
MAX_UNCOMPRESSED_BYTES = 200 * 1024 * 1024

#: Rows in one sheet. A guard against a workbook whose used range runs to the format's million,
#: which openpyxl will faithfully iterate over.
MAX_ROWS = 50_000

MACRO_MEMBERS = ("vbaproject.bin", "xl/macrosheets/", "xl/vbaproject.bin")


class UnreadableFile(ValueError):
    """The upload is not something this importer will open, and why."""


@dataclasses.dataclass(frozen=True)
class Sheet:
    """One sheet, read into memory.

    Read eagerly rather than streamed: the dry run visits every row twice — once to classify and
    once to look for duplicates within the file — and a read-only openpyxl worksheet is a
    one-pass iterator that silently yields nothing the second time.
    """

    name: str
    headings: list[str]
    #: ``(row_number_in_the_file, {heading: value})``. The number is the file's, so a message
    #: about row 214 can be acted on by opening the file and going to row 214.
    rows: list[tuple[int, dict[str, Any]]]


@dataclasses.dataclass(frozen=True)
class ParsedFile:
    content_sha256: str
    size: int
    sheets: list[Sheet]

    def sheet(self, name: str) -> Sheet | None:
        return next((sheet for sheet in self.sheets if sheet.name == name), None)


def digest(content: bytes) -> str:
    """The SHA-256 §17.1 asks a commit to verify against its dry run."""
    return hashlib.sha256(content).hexdigest()


def read(content: bytes, *, sheet_names: tuple[str, ...] | None = None) -> ParsedFile:
    """Open an uploaded workbook and hand back its cells.

    ``sheet_names`` narrows the read. The export writes three sheets and only one holds data;
    reading the Data Dictionary as though it were allocations would produce a screen full of
    error rows about a sheet that is doing exactly what it should.
    """
    _check_size(content)
    _check_archive(content)

    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            # The whole point (§17.1). A formula cell yields its formula, which is refused, and
            # never the value some other program cached for it.
            data_only=False,
            keep_links=False,
        )
    except UnreadableFile:
        raise
    except Exception as exc:  # openpyxl raises a wide variety for a malformed file
        raise UnreadableFile(f"This file could not be opened as a spreadsheet: {exc}") from exc

    try:
        wanted = sheet_names or tuple(workbook.sheetnames)
        sheets = [
            _read_sheet(workbook[name]) for name in workbook.sheetnames if name in set(wanted)
        ]
    finally:
        # A read-only workbook holds the archive open, and an unclosed ZipFile surfaces as an
        # unraisable exception during garbage collection rather than as a failure here.
        workbook.close()

    return ParsedFile(content_sha256=digest(content), size=len(content), sheets=sheets)


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------
def _check_size(content: bytes) -> None:
    if not content:
        raise UnreadableFile("The uploaded file is empty.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise UnreadableFile(
            f"The uploaded file is {len(content)} bytes, over the "
            f"{MAX_UPLOAD_BYTES}-byte limit for an import."
        )


def _check_archive(content: bytes) -> None:
    """Everything that can be decided from the container alone.

    Done against the central directory rather than by extracting, because the entire value of
    a zip-bomb check is that it happens before anything is expanded.
    """
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile as exc:
        raise UnreadableFile(
            "This is not an .xlsx workbook. An .xlsx file is a ZIP archive, and this one is "
            "not — a .xls or .csv saved with the wrong extension looks like this."
        ) from exc

    with archive:
        names = archive.namelist()
        if "xl/workbook.xml" not in names:
            raise UnreadableFile(
                "This ZIP archive is not a spreadsheet: it has no xl/workbook.xml."
            )

        lowered = [name.lower() for name in names]
        if any(name.startswith(MACRO_MEMBERS) or name.endswith(MACRO_MEMBERS) for name in lowered):
            raise UnreadableFile(
                "This workbook contains macros. Save it as .xlsx — a macro-enabled workbook is "
                "not accepted for an import, whether or not the macros would run."
            )

        declared = sum(info.file_size for info in archive.infolist())
        if declared > MAX_UNCOMPRESSED_BYTES:
            raise UnreadableFile(
                f"This archive declares {declared} bytes of content, over the "
                f"{MAX_UNCOMPRESSED_BYTES}-byte limit. It was not opened."
            )


def _read_sheet(sheet: Any) -> Sheet:
    """One sheet's headings and rows, with blank rows dropped."""
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return Sheet(name=sheet.title, headings=[], rows=[])

    headings = [_heading_text(value) for value in header]
    rows: list[tuple[int, dict[str, Any]]] = []

    for offset, values in enumerate(rows_iter, start=2):
        if offset > MAX_ROWS:
            raise UnreadableFile(
                f"Sheet {sheet.title!r} has more than {MAX_ROWS} rows. Split the file."
            )
        if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
            # A trailing blank row is how most spreadsheets end and is not a row about anything.
            continue
        rows.append(
            (
                offset,
                {
                    heading: _cell(value)
                    for heading, value in zip(headings, values, strict=False)
                    if heading
                },
            )
        )
    return Sheet(name=sheet.title, headings=[h for h in headings if h], rows=rows)


def _heading_text(value: Any) -> str:
    """A heading as text, with the export's own formula guard removed.

    An export of a column whose heading began with a formula character would have written it
    guarded; reading it back has to undo that, or the file the platform produced would not match
    the vocabulary the platform reads.
    """
    if value is None:
        return ""
    return str(restore(value)).strip()


def _cell(value: Any) -> Any:
    """One cell, exactly as stored, with surrounding space removed.

    §21.12's apostrophe is deliberately **not** undone here. It is the one thing that tells a
    guarded literal apart from a real formula: a Satnet Path code of ``-Ka`` was written as
    ``'-Ka`` and is data, while ``=SUM(A1)`` is a formula and is refused. Undoing the guard
    before that decision is made would turn every guarded code into a formula and refuse the
    export this platform produced. :mod:`normalize` undoes it, after deciding.
    """
    if isinstance(value, str):
        return value.strip()
    return value
