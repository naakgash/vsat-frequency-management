"""Writing a workbook nobody has to trust. §17.2, §21.12, §26.19.

Every cell in every export goes through :func:`write_row`, which is the only place a value
reaches openpyxl. That is the point: formula neutralisation is not something each sheet
remembers to do, it is the only way to write a cell at all.

Three sheets, always, and the last two are what make an export auditable rather than merely
useful:

* **the data** — whatever the caller is exporting;
* **Data Dictionary** — what every column means, taken from the Specification Dictionary rather
  than restated here (§2);
* **Export** — who ran it, when, with which filters and which columns, and how many rows came
  back. §17.2 asks for the filter parameters to be recorded, and the reason is that an export
  without them is a number nobody can reproduce.
"""

from __future__ import annotations

import dataclasses
import datetime
import io
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from imports_exports.export.safety import neutralise

HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

#: How wide a column is allowed to get from its content. Excel's own auto-fit is not available
#: through openpyxl, and a column sized to its widest cell makes a Satnet Path code column three
#: screens wide the first time somebody writes a long note.
MAX_COLUMN_WIDTH = 60


@dataclasses.dataclass(frozen=True)
class Provenance:
    """What the Export sheet records about a run. §17.2."""

    exported_by: str
    exported_at: datetime.datetime
    source: str
    filters: dict[str, str]
    columns: list[str]
    row_count: int
    notes: str = ""


def new_workbook() -> Workbook:
    """A workbook with no default sheet, so every sheet in it was asked for."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def write_header(sheet: Worksheet, headings: list[str]) -> None:
    for index, heading in enumerate(headings, start=1):
        cell = sheet.cell(row=1, column=index, value=neutralise(heading))
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
    sheet.freeze_panes = "A2"


def write_row(sheet: Worksheet, row_index: int, values: list[Any]) -> None:
    """Write one row, neutralising every string on the way in.

    The single choke point. A sheet that built cells itself would be one refactor away from
    forgetting §21.12, and the omission would be invisible until somebody opened the file.
    """
    for index, value in enumerate(values, start=1):
        sheet.cell(row=row_index, column=index, value=neutralise(_writable(value)))


def _writable(value: Any) -> Any:
    """Convert a value openpyxl cannot store into one it can.

    **Timestamps lose their tzinfo, and that is a real collision with A-28.** The xlsx format
    has no concept of a time zone and openpyxl refuses an aware datetime outright, so there are
    two ways out: write the value as an ISO string with its offset, or convert to UTC and write
    a true datetime cell. This takes the second — a text timestamp cannot be sorted or filtered
    as a date, which is most of what somebody opens a spreadsheet to do.

    The zone does not go unstated: it moves to the **column heading**, which is what
    `normalized._heading` appends "(UTC)" for. That keeps A-28's rule — a timestamp always says
    which zone it is in — while letting the cell stay a date.

    UUIDs become their canonical string, which is what lets an export come back through the S15
    importer. A Decimal is left alone: openpyxl stores it exactly, and converting it to float
    here would undo ADR-0003 at the last possible moment.
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.astimezone(datetime.UTC).replace(tzinfo=None)
    if isinstance(value, bool | int | float | str | datetime.date):
        return value
    if isinstance(value, Decimal):
        return value
    return str(value)


def autosize(sheet: Worksheet, headings: list[str]) -> None:
    """Size each column to its content, within a bound."""
    for index, heading in enumerate(headings, start=1):
        longest = len(str(heading))
        for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
            for cell in row:
                longest = max(longest, len(str(cell.value or "")))
        sheet.column_dimensions[get_column_letter(index)].width = min(longest + 2, MAX_COLUMN_WIDTH)


def add_data_dictionary(workbook: Workbook, codes: list[str]) -> Worksheet:
    """One sheet explaining every specification code the export used. §2, §26.19.

    Read from the dictionary, never restated: an export that carried its own descriptions would
    be a second copy of the thing §2 exists to keep singular, and it would go stale the first
    time an administrator improved a description.
    """
    from specifications.models import SpecificationDefinition

    sheet = workbook.create_sheet("Data Dictionary")
    headings = ["Code", "Name", "Unit", "Data type", "Description", "Calculation"]
    write_header(sheet, headings)

    definitions = {
        definition.code: definition
        for definition in SpecificationDefinition.objects.filter(code__in=codes)
    }
    for row_index, code in enumerate(sorted(codes), start=2):
        definition = definitions.get(code)
        write_row(
            sheet,
            row_index,
            [
                code,
                definition.display_name if definition else "",
                definition.unit if definition else "",
                definition.data_type if definition else "",
                definition.description if definition else "",
                definition.calculation_note if definition else "",
            ],
        )
    autosize(sheet, headings)
    return sheet


def add_provenance(workbook: Workbook, provenance: Provenance) -> Worksheet:
    """The sheet that makes the numbers reproducible. §17.2.

    Timestamps are UTC and say so (**A-28**): a workbook is read somewhere else, later, by
    somebody who was not there when it was produced.
    """
    sheet = workbook.create_sheet("Export")
    headings = ["Field", "Value"]
    write_header(sheet, headings)

    rows: list[tuple[str, Any]] = [
        ("Exported by", provenance.exported_by),
        ("Exported at (UTC)", provenance.exported_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Source", provenance.source),
        ("Rows", provenance.row_count),
        ("Columns", ", ".join(provenance.columns)),
    ]
    rows += [(f"Filter: {key}", value) for key, value in sorted(provenance.filters.items())] or [
        ("Filter", "none — every row within the exporter's scope")
    ]
    if provenance.notes:
        rows.append(("Notes", provenance.notes))

    for row_index, (field, value) in enumerate(rows, start=2):
        write_row(sheet, row_index, [field, value])
    autosize(sheet, headings)
    return sheet


def to_bytes(workbook: Workbook) -> bytes:
    """Serialise in memory. An export is a download, not a file on the application's disk."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
