"""The two-stage import. §17.1, §26.17, §26.19, ADR-0015.

Three properties this file exists to hold, and everything else is detail:

**Nothing a spreadsheet calculated is ever used.** A formula is refused rather than evaluated,
the cached answer openpyxl could have handed over is never asked for, and a derived column that
disagrees with the engine produces a warning and the *engine's* number. The test that matters
most is the last kind: the file says one occupied bandwidth, the platform stores another, and
the row commits.

**A free-capacity row is never an allocation.** §17.1 says ignore them; ADR-0009 says why they
can never be anything else — free capacity is computed from what is allocated, so importing a
gap row would reserve the very spectrum it says is free.

**What commits is what was reviewed.** The SHA-256 proves the file has not changed; writing from
the stored rows proves the numbers on the screen are the numbers that land.
"""

from __future__ import annotations

import datetime
import io
import uuid
import zipfile
from typing import Any

import pytest
from django.core.exceptions import PermissionDenied
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from accounts.constants import Role
from audit.models import AuditEvent
from imports_exports import services
from imports_exports.constants import BatchPolicy, ImportStage, RowClassification
from imports_exports.export import normalized
from imports_exports.export import workbook as writer
from imports_exports.importer import fields as field_registry
from imports_exports.importer import mapping, parse
from imports_exports.models import ImportBatch, ImportMapping
from satnet_paths.constants import InputMode, PathStatus
from satnet_paths.models import SatnetPath
from spectrum.models import SpectrumReservation
from tests.factories import make_user

pytestmark = pytest.mark.django_db

MHZ = 1_000_000

H = {field.key: field.heading for field in field_registry.INPUT_FIELDS}


# ---------------------------------------------------------------------------
# Building files to feed it
# ---------------------------------------------------------------------------
def build_workbook(
    rows: list[dict[str, Any]],
    *,
    sheet_name: str = normalized.SHEET_NAME,
    headings: list[str] | None = None,
    extra_sheets: dict[str, list[str]] | None = None,
) -> bytes:
    """A workbook written the way a person's spreadsheet is: values straight into cells.

    Deliberately **not** built through `export.workbook`. An importer tested only against files
    its own exporter produced is an importer tested against its own assumptions, and §21.12's
    guard would hide every dangerous value before the importer ever saw one.
    """
    book = Workbook()
    book.remove(book.active)
    sheet = book.create_sheet(sheet_name)

    columns = headings or sorted({key for row in rows for key in row}) or list(H.values())
    for index, heading in enumerate(columns, start=1):
        sheet.cell(row=1, column=index, value=heading)
    for offset, row in enumerate(rows, start=2):
        for index, heading in enumerate(columns, start=1):
            sheet.cell(row=offset, column=index, value=row.get(heading))

    for name, values in (extra_sheets or {}).items():
        other = book.create_sheet(name)
        for index, value in enumerate(values, start=1):
            other.cell(row=1, column=index, value=value)

    buffer = io.BytesIO()
    book.save(buffer)
    book.close()
    return buffer.getvalue()


def a_row(world, **overrides: Any) -> dict[str, Any]:
    """One good allocation, as cells keyed by the heading the export writes."""
    row = {
        H["code"]: "IMP-1",
        H["satnet"]: world["satnet"].code,
        H["direction"]: "FWD",
        H["input_mode"]: InputMode.OCCUPIED_BW,
        H["input_value"]: 10 * MHZ,
        H["rolloff"]: "0.2",
        H["canonical_center_hz"]: 50 * MHZ,
        H["valid_from"]: world["valid_from"],
    }
    row.update(overrides)
    return row


@pytest.fixture
def world(lifecycle_world):
    """The lifecycle world, plus an administrator and a timestamp inside every validity period.

    The moment is derived from the clock rather than written out: the Satnet is effective from
    yesterday, so a fixed date expires the day the calendar passes it and the failure looks like
    an import bug rather than a stale fixture.
    """
    return {
        **lifecycle_world,
        "importer": lifecycle_world["admin"],
        "valid_from": (timezone.now() + timezone.timedelta(hours=1)).replace(
            second=0, microsecond=0, tzinfo=None
        ),
    }


def dry_run(world, rows, **kwargs: Any) -> ImportBatch:
    # `is None` rather than `or`: an empty upload is one of the things under test, and empty
    # bytes are falsy.
    content = kwargs.pop("content", None)
    content = build_workbook(rows) if content is None else content
    return services.dry_run(
        actor=kwargs.pop("actor", world["importer"]),
        content=content,
        file_name=kwargs.pop("file_name", "plan.xlsx"),
        **kwargs,
    )


def classification_of(batch: ImportBatch, row_number: int = 2) -> str:
    return batch.rows.get(row_number=row_number).classification


# ---------------------------------------------------------------------------
# §21 — what is refused before openpyxl is asked to read anything
# ---------------------------------------------------------------------------
def test_something_that_is_not_a_zip_is_refused_with_a_reason(world):
    with pytest.raises(parse.UnreadableFile) as caught:
        dry_run(world, [], content=b"code,satnet\nIMP-1,SN-LC\n")

    assert "ZIP" in str(caught.value)


def test_an_empty_upload_is_refused(world):
    with pytest.raises(parse.UnreadableFile):
        dry_run(world, [], content=b"")


def test_an_oversized_upload_is_refused_without_being_opened(world, monkeypatch):
    monkeypatch.setattr(parse, "MAX_UPLOAD_BYTES", 10)

    with pytest.raises(parse.UnreadableFile) as caught:
        dry_run(world, [a_row(world)])

    assert "limit" in str(caught.value)


def test_a_macro_enabled_workbook_is_refused(world):
    """openpyxl runs nothing, and the file is refused anyway.

    A workbook carrying `vbaProject.bin` has no business in a frequency plan, and refusing it
    means nobody has to reason about what a later library version might do with it.
    """
    content = build_workbook([a_row(world)])
    buffer = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content)) as source, zipfile.ZipFile(buffer, "w") as target:
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr("xl/vbaProject.bin", b"\x00\x01")

    with pytest.raises(parse.UnreadableFile) as caught:
        dry_run(world, [], content=buffer.getvalue())

    assert "macros" in str(caught.value)


def test_an_archive_claiming_an_absurd_expansion_is_refused(world, monkeypatch):
    monkeypatch.setattr(parse, "MAX_UNCOMPRESSED_BYTES", 1)

    with pytest.raises(parse.UnreadableFile) as caught:
        dry_run(world, [a_row(world)])

    assert "not opened" in str(caught.value)


def test_a_file_with_no_recognised_column_is_refused_rather_than_read_as_errors(world):
    content = build_workbook([{"Frequency": 1, "Notes": "x"}])

    with pytest.raises(parse.UnreadableFile) as caught:
        dry_run(world, [], content=content)

    assert H["code"] in str(caught.value)


def test_the_exports_own_extra_sheets_are_skipped(world):
    """A Data Dictionary sheet is doing its job; reading it as allocations would be noise."""
    content = build_workbook(
        [a_row(world)],
        extra_sheets={writer.DATA_DICTIONARY_SHEET: ["Code", "Name"], writer.PROVENANCE_SHEET: []},
    )
    batch = dry_run(world, [], content=content)

    assert {row.sheet for row in batch.rows.all()} == {normalized.SHEET_NAME}


# ---------------------------------------------------------------------------
# §17.1 — an Excel-calculated value is never used
# ---------------------------------------------------------------------------
def test_a_formula_in_an_input_cell_is_refused_and_never_evaluated(world):
    batch = dry_run(world, [a_row(world, **{H["input_value"]: "=10*1000000"})])
    row = batch.rows.get(row_number=2)

    assert row.classification == RowClassification.ERROR
    codes = {message["code"] for message in row.messages}
    assert "FORMULA" in codes
    # The formula is quoted back, not resolved to the number it would produce.
    assert any("=10*1000000" in message["text"] for message in row.messages)


def test_the_workbook_is_never_opened_for_its_cached_values(world, monkeypatch):
    """The structural half of §17.1, and the one no data fixture can express.

    ``data_only=True`` would hand over the number some other program computed, at some other
    time, possibly against inputs that have since changed. That number is precisely what this
    platform exists not to depend on, so the flag is asserted rather than assumed.
    """
    seen: dict[str, Any] = {}
    original = parse.load_workbook

    def record(*args: Any, **kwargs: Any):
        seen.update(kwargs)
        return original(*args, **kwargs)

    monkeypatch.setattr(parse, "load_workbook", record)
    dry_run(world, [a_row(world)])

    assert seen["data_only"] is False
    assert seen["keep_links"] is False


def test_a_derived_column_the_file_disagrees_about_is_a_warning_and_the_engine_wins(world):
    """The heart of §17.1. The spreadsheet says one bandwidth; the platform stores its own."""
    content = build_workbook([a_row(world, **{_derived_heading("occupied_bw"): 9 * MHZ})])
    batch = dry_run(world, [], content=content)
    row = batch.rows.get(row_number=2)

    assert row.classification == RowClassification.WARNING
    assert any(message["code"] == "RECALCULATED" for message in row.messages)

    services.commit_batch(actor=world["importer"], batch=batch, content=content)
    assert SatnetPath.objects.get().occupied_bw_hz == 10 * MHZ


def test_a_frequency_carrying_its_own_unit_is_refused_rather_than_guessed_at(world):
    """A-08 and §26.20. Being wrong by a factor of a million looks entirely plausible."""
    batch = dry_run(world, [a_row(world, **{H["canonical_center_hz"]: "50 MHz"})])

    assert classification_of(batch) == RowClassification.ERROR


def test_a_fraction_of_a_hertz_is_refused(world):
    batch = dry_run(world, [a_row(world, **{H["canonical_center_hz"]: 50_000_000.5})])

    assert classification_of(batch) == RowClassification.ERROR


def test_a_thousands_separator_is_forgiven(world):
    """A spreadsheet shows one whether or not it stored one."""
    batch = dry_run(world, [a_row(world, **{H["input_value"]: "10,000,000"})])

    assert classification_of(batch) == RowClassification.VALID
    assert batch.rows.get(row_number=2).normalized["input_value"] == 10 * MHZ


def test_a_naive_timestamp_is_read_as_utc(world):
    """A-28, ADR-0022. The xlsx format has no zone, and the heading says which one to assume."""
    batch = dry_run(world, [a_row(world)])
    stored = batch.rows.get(row_number=2).normalized["valid_from"]

    assert datetime.datetime.fromisoformat(stored) == world["valid_from"].replace(
        tzinfo=datetime.UTC
    )


def test_a_value_guarded_against_formula_injection_comes_back_as_it_went_out(world):
    """§21.12's apostrophe is an export concern; an import has to undo exactly it and no more.

    Without this, every Satnet Path code beginning with a formula character would gain a
    character on each round trip and stop matching the record it came from.
    """
    batch = dry_run(world, [a_row(world, **{H["code"]: "'-Ka-1"})])
    row = batch.rows.get(row_number=2)

    assert row.classification == RowClassification.VALID
    assert row.normalized["code"] == "-Ka-1"


def test_a_real_formula_is_still_refused_after_the_guard_is_understood(world):
    """The other half of the pair above: `'=x` is data and `=x` is a formula."""
    batch = dry_run(world, [a_row(world, **{H["code"]: "=SUM(A1:A2)"})])

    assert classification_of(batch) == RowClassification.ERROR


# ---------------------------------------------------------------------------
# The seven classifications
# ---------------------------------------------------------------------------
def test_a_good_row_is_valid(world):
    batch = dry_run(world, [a_row(world)])

    assert classification_of(batch) == RowClassification.VALID
    assert batch.counts[RowClassification.VALID] == 1


def test_a_missing_required_cell_is_an_error(world):
    batch = dry_run(world, [a_row(world, **{H["rolloff"]: None})])
    row = batch.rows.get(row_number=2)

    assert row.classification == RowClassification.ERROR
    assert {message["code"] for message in row.messages} == {"MISSING"}


def test_a_label_naming_nothing_needs_a_mapping(world):
    batch = dry_run(world, [a_row(world, **{H["satnet"]: "Ka Hub 1 (old)"})])
    row = batch.rows.get(row_number=2)

    assert row.classification == RowClassification.NEEDS_MAPPING
    assert "Ka Hub 1 (old)" in row.messages[0]["text"]


def test_an_allocation_already_here_is_a_duplicate_rather_than_a_second_copy(world, make_path):
    make_path(PathStatus.DRAFT, code="IMP-1")
    batch = dry_run(world, [a_row(world)])
    row = batch.rows.get(row_number=2)

    assert row.classification == RowClassification.DUPLICATE
    assert row.resulting_object_id == SatnetPath.objects.get().pk


def test_the_same_allocation_twice_in_one_file_is_a_duplicate(world):
    batch = dry_run(world, [a_row(world), a_row(world, **{H["canonical_center_hz"]: 80 * MHZ})])

    assert classification_of(batch, 2) == RowClassification.VALID
    assert classification_of(batch, 3) == RowClassification.DUPLICATE


def test_a_row_whose_spectrum_is_taken_is_a_conflict(world, make_path):
    make_path(PathStatus.ON_AIR, code="LC-HOLD", centre=50 * MHZ)
    batch = dry_run(world, [a_row(world)])
    row = batch.rows.get(row_number=2)

    assert row.classification == RowClassification.CONFLICT
    assert any("overlaps" in message["text"] for message in row.messages)


@pytest.mark.parametrize("marker", ["Free", "FREE CAPACITY", "spare", "unallocated", "Gap"])
def test_a_free_capacity_row_is_ignored_however_it_says_so(world, marker):
    batch = dry_run(world, [a_row(world, **{H["code"]: marker})])

    assert classification_of(batch) == RowClassification.IGNORED_FREE_CAPACITY


def test_a_row_carrying_a_frequency_and_naming_nothing_is_free_capacity(world):
    """The way a hand-maintained sheet usually writes a gap: a range and empty identity cells."""
    batch = dry_run(world, [a_row(world, **{H["code"]: None, H["satnet"]: None})])

    assert classification_of(batch) == RowClassification.IGNORED_FREE_CAPACITY


def test_a_free_capacity_row_never_becomes_a_satnet_path(world):
    """§17.1 and ADR-0009 together — the property, not the classification.

    Importing a gap row as an allocation would reserve the spectrum it says is free, and the
    free-capacity engine would then stop reporting it. This asserts the outcome after a commit
    rather than the label at the dry run, because the label is not what protects anything.
    """
    content = build_workbook([a_row(world), a_row(world, **{H["code"]: "FREE"})])
    batch = dry_run(world, [], content=content)
    services.commit_batch(actor=world["importer"], batch=batch, content=content)

    assert list(SatnetPath.objects.values_list("code", flat=True)) == ["IMP-1"]
    assert batch.rows.get(row_number=3).resulting_object_id is None


def test_every_classification_is_reachable(world, make_path):
    """All seven, from one file. The count is the plan's acceptance criterion, not a summary."""
    make_path(PathStatus.ON_AIR, code="LC-HOLD", centre=90 * MHZ)
    make_path(PathStatus.DRAFT, code="IMP-DUP", centre=20 * MHZ)

    batch = dry_run(
        world,
        [
            a_row(world, **{H["code"]: "IMP-OK", H["canonical_center_hz"]: 50 * MHZ}),
            a_row(
                world,
                **{
                    H["code"]: "IMP-WARN",
                    H["canonical_center_hz"]: 65 * MHZ,
                    _derived_heading("occupied_bw"): 9 * MHZ,
                },
            ),
            a_row(world, **{H["code"]: "IMP-BAD", H["rolloff"]: None}),
            a_row(world, **{H["code"]: "IMP-MAP", H["satnet"]: "Nothing"}),
            a_row(world, **{H["code"]: "IMP-DUP"}),
            a_row(world, **{H["code"]: "IMP-CLASH", H["canonical_center_hz"]: 90 * MHZ}),
            a_row(world, **{H["code"]: "FREE"}),
        ],
    )

    assert {value: count for value, count in batch.counts.items() if count} == {
        RowClassification.VALID: 1,
        RowClassification.WARNING: 1,
        RowClassification.ERROR: 1,
        RowClassification.NEEDS_MAPPING: 1,
        RowClassification.DUPLICATE: 1,
        RowClassification.CONFLICT: 1,
        RowClassification.IGNORED_FREE_CAPACITY: 1,
    }


# ---------------------------------------------------------------------------
# Stage one — the dry run touches no production data
# ---------------------------------------------------------------------------
def test_a_dry_run_writes_no_allocation(world):
    dry_run(world, [a_row(world), a_row(world, **{H["code"]: "IMP-2"})])

    assert not SatnetPath.objects.exists()
    assert not SpectrumReservation.objects.exists()


def test_a_dry_run_records_the_file_and_its_digest(world):
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content, file_name="february.xlsx")

    assert batch.file_name == "february.xlsx"
    assert batch.file_sha256 == parse.digest(content)
    assert batch.stage == ImportStage.DRY_RUN
    assert batch.row_count == 1


def test_a_dry_run_is_audited(world):
    batch = dry_run(world, [a_row(world)])
    event = AuditEvent.objects.get(action="IMPORT_DRY_RUN")

    assert event.import_batch_id == batch.pk
    assert event.after["sha256"] == batch.file_sha256


def test_the_row_number_is_the_files_own(world):
    """A message about row 4 has to mean row 4 of the spreadsheet, blank rows and all."""
    batch = dry_run(world, [a_row(world), a_row(world, **{H["code"]: "IMP-2"})])

    assert sorted(batch.rows.values_list("row_number", flat=True)) == [2, 3]


# ---------------------------------------------------------------------------
# Stage two — the SHA-256 seam
# ---------------------------------------------------------------------------
def test_a_commit_refuses_a_file_that_has_changed_since_it_was_read(world):
    batch = dry_run(world, [a_row(world)])
    edited = build_workbook([a_row(world, **{H["canonical_center_hz"]: 80 * MHZ})])

    with pytest.raises(services.CommitRefused) as caught:
        services.commit_batch(actor=world["importer"], batch=batch, content=edited)

    assert "does not match" in str(caught.value)
    assert not SatnetPath.objects.exists()
    batch.refresh_from_db()
    assert batch.stage == ImportStage.DRY_RUN


def test_a_refused_commit_is_audited(world):
    batch = dry_run(world, [a_row(world)])

    with pytest.raises(services.CommitRefused):
        services.commit_batch(actor=world["importer"], batch=batch, content=b"PK\x03\x04nonsense")

    assert AuditEvent.objects.filter(action="IMPORT_REFUSED").exists()


def test_the_same_file_commits(world):
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content)

    services.commit_batch(actor=world["importer"], batch=batch, content=content)
    batch.refresh_from_db()

    assert batch.stage == ImportStage.COMMITTED
    assert SatnetPath.objects.get().code == "IMP-1"


def test_committing_twice_is_refused_rather_than_repeated(world):
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content)
    services.commit_batch(actor=world["importer"], batch=batch, content=content)
    batch.refresh_from_db()

    with pytest.raises(services.CommitRefused) as caught:
        services.commit_batch(actor=world["importer"], batch=batch, content=content)

    assert "already committed" in str(caught.value)
    assert SatnetPath.objects.count() == 1


def test_a_committed_row_points_at_what_it_produced(world):
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content)
    services.commit_batch(actor=world["importer"], batch=batch, content=content)

    assert batch.rows.get(row_number=2).resulting_object_id == SatnetPath.objects.get().pk


def test_a_commit_is_audited(world):
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content)
    services.commit_batch(actor=world["importer"], batch=batch, content=content, reason="cutover")

    event = AuditEvent.objects.get(action="IMPORT_COMMITTED")
    assert event.after["created"] == 1
    assert event.change_reason == "cutover"


# ---------------------------------------------------------------------------
# What an import is allowed to produce
# ---------------------------------------------------------------------------
def test_everything_imported_is_a_draft(world):
    """An import is bulk data entry, not an approval (§12, §15.2)."""
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content)
    services.commit_batch(actor=world["importer"], batch=batch, content=content)

    assert SatnetPath.objects.get().status == PathStatus.DRAFT


def test_an_imported_conflict_is_reported_and_holds_no_spectrum(world, make_path):
    """§17.1's "reported and not activated", as a property rather than a label.

    The row is carried across — losing it would hide the overlap the migration exists to
    surface — and it reserves nothing, because a draft holds no spectrum.
    """
    make_path(PathStatus.ON_AIR, code="LC-HOLD", centre=50 * MHZ)
    before = SpectrumReservation.objects.count()

    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content, batch_policy=BatchPolicy.ROW_BY_ROW)
    services.commit_batch(actor=world["importer"], batch=batch, content=content)

    imported = SatnetPath.objects.get(code="IMP-1")
    assert imported.status == PathStatus.DRAFT
    assert SpectrumReservation.objects.count() == before
    assert batch.rows.get(row_number=2).classification == RowClassification.CONFLICT


def test_an_identifier_from_an_export_is_honoured(world):
    """§17.1's stable UUIDs: a re-import is the same allocation, not a copy of it."""
    identifier = uuid.uuid4()
    group = uuid.uuid4()
    content = build_workbook(
        [{**a_row(world), "id": str(identifier), "revision_group": str(group)}]
    )
    batch = dry_run(world, [], content=content)
    services.commit_batch(actor=world["importer"], batch=batch, content=content)

    path = SatnetPath.objects.get()
    assert path.pk == identifier
    assert path.revision_group == group


def test_an_unreadable_identifier_is_an_error_rather_than_ignored(world):
    """Ignoring it would create a second allocation beside the one it meant to be."""
    batch = dry_run(world, [{**a_row(world), "id": "not-a-uuid"}])
    row = batch.rows.get(row_number=2)

    assert row.classification == RowClassification.ERROR
    assert any(message["code"] == "BAD_IDENTIFIER" for message in row.messages)


# ---------------------------------------------------------------------------
# The two batch policies — docs/design/04 §8.4
# ---------------------------------------------------------------------------
def test_all_or_nothing_refuses_a_batch_holding_a_blocking_row(world):
    content = build_workbook([a_row(world), a_row(world, **{H["code"]: None, H["rolloff"]: None})])
    batch = dry_run(world, [], content=content, batch_policy=BatchPolicy.ALL_OR_NOTHING)

    with pytest.raises(services.CommitRefused) as caught:
        services.commit_batch(actor=world["importer"], batch=batch, content=content)

    assert "all or nothing" in str(caught.value)
    assert not SatnetPath.objects.exists()


def test_row_by_row_keeps_what_worked(world):
    content = build_workbook(
        [
            a_row(world, **{H["code"]: "IMP-OK"}),
            a_row(world, **{H["code"]: "IMP-BAD", H["rolloff"]: "nonsense"}),
            a_row(world, **{H["code"]: "IMP-ALSO-OK", H["canonical_center_hz"]: 80 * MHZ}),
        ]
    )
    batch = dry_run(world, [], content=content, batch_policy=BatchPolicy.ROW_BY_ROW)
    services.commit_batch(actor=world["importer"], batch=batch, content=content)

    assert set(SatnetPath.objects.values_list("code", flat=True)) == {"IMP-OK", "IMP-ALSO-OK"}


def test_a_write_that_fails_late_is_recorded_rather_than_raised(world, monkeypatch):
    """A row that survived classification and is refused by the write.

    Forced rather than contrived from data: the point is the *policy's* behaviour when the
    services refuse something classification could not have seen, and a reviewer who gets a
    stack trace instead of a row learns nothing about which row it was.
    """
    from django.core.exceptions import ValidationError

    from imports_exports.importer import commit as commit_stage

    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content, batch_policy=BatchPolicy.ROW_BY_ROW)
    monkeypatch.setattr(
        commit_stage.path_services,
        "create",
        lambda **_: (_ for _ in ()).throw(ValidationError("refused late")),
    )

    services.commit_batch(actor=world["importer"], batch=batch, content=content)

    row = batch.rows.get(row_number=2)
    assert row.classification == RowClassification.ERROR
    assert any(message["code"] == "WRITE_FAILED" for message in row.messages)
    assert not SatnetPath.objects.exists()


def test_a_late_failure_stops_an_all_or_nothing_batch_with_a_sentence(world, monkeypatch):
    from django.core.exceptions import ValidationError

    from imports_exports.importer import commit as commit_stage

    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content, batch_policy=BatchPolicy.ALL_OR_NOTHING)
    monkeypatch.setattr(
        commit_stage.path_services,
        "create",
        lambda **_: (_ for _ in ()).throw(ValidationError("refused late")),
    )

    with pytest.raises(services.CommitRefused) as caught:
        services.commit_batch(actor=world["importer"], batch=batch, content=content)

    assert "all or nothing" in str(caught.value)
    assert not SatnetPath.objects.exists()


def test_an_unrecognised_policy_falls_back_to_the_cautious_one(world, client):
    """The value arrives from a form, and the safe default is the one where a surprise stops."""
    client.force_login(world["importer"])
    client.post(
        reverse("imports:dry-run"),
        {
            "file": _uploaded(build_workbook([a_row(world)])),
            "batch_policy": "WHATEVER",
        },
    )

    assert ImportBatch.objects.get().batch_policy == BatchPolicy.ALL_OR_NOTHING


# ---------------------------------------------------------------------------
# Re-checking on commit — §9.5's rule, applied to a batch
# ---------------------------------------------------------------------------
def test_a_row_whose_spectrum_was_taken_after_the_review_is_reclassified(world, make_path):
    """A review on Monday and a commit on Wednesday are two different worlds."""
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content, batch_policy=BatchPolicy.ROW_BY_ROW)
    assert batch.rows.get(row_number=2).classification == RowClassification.VALID

    make_path(PathStatus.ON_AIR, code="LC-LATE", centre=50 * MHZ)
    services.commit_batch(actor=world["importer"], batch=batch, content=content)

    row = batch.rows.get(row_number=2)
    assert row.classification == RowClassification.CONFLICT
    assert any(message["code"] == "RECLASSIFIED" for message in row.messages)


# ---------------------------------------------------------------------------
# Remembered mappings
# ---------------------------------------------------------------------------
def test_a_remembered_label_resolves_on_the_next_run(world):
    rows = [a_row(world, **{H["satnet"]: "Ka Hub 1 (old)"})]
    assert classification_of(dry_run(world, rows)) == RowClassification.NEEDS_MAPPING

    mapping.remember(
        actor=world["importer"],
        kind=field_registry.SATNET,
        label="Ka Hub 1 (old)",
        target_id=world["satnet"].pk,
    )

    assert classification_of(dry_run(world, rows)) == RowClassification.VALID


def test_remembering_a_mapping_is_audited(world):
    mapping.remember(
        actor=world["importer"],
        kind=field_registry.SATNET,
        label="Ka Hub 1 (old)",
        target_id=world["satnet"].pk,
    )

    assert AuditEvent.objects.filter(action="IMPORT_MAPPING_REMEMBERED").exists()
    assert ImportMapping.objects.get().target_id == world["satnet"].pk


def test_a_near_match_is_never_resolved_on_its_own(world):
    """The one mistake an import must not make unsupervised."""
    resolution = mapping.resolve(field_registry.SATNET, f"{world['satnet'].code}-X")

    assert not resolution.ok


def test_a_code_resolves_regardless_of_case(world):
    resolution = mapping.resolve(field_registry.SATNET, world["satnet"].code.lower())

    assert resolution.ok
    assert resolution.source == "code"


def test_an_unknown_label_is_asked_about_once_however_many_rows_use_it(world):
    batch = dry_run(
        world,
        [
            a_row(world, **{H["code"]: f"IMP-{index}", H["satnet"]: "Ka Hub 1 (old)"})
            for index in range(5)
        ],
    )

    from imports_exports import selectors

    assert len(selectors.unresolved_labels(batch)) == 1


# ---------------------------------------------------------------------------
# The round trip — the export is the shape the import reads
# ---------------------------------------------------------------------------
def test_the_exports_headings_are_what_the_importer_expects(world, make_path):
    """One chain: dictionary → column registry → export heading → importer field.

    A renamed column has to break here, visibly, rather than producing a file the product that
    wrote it can no longer read.
    """
    make_path(PathStatus.DRAFT, code="RT-1")
    export = services.export_satnet_paths(
        actor=world["importer"], columns=[field.column for field in field_registry.INPUT_FIELDS]
    )
    written = parse.read(export.content).sheet(normalized.SHEET_NAME)

    assert written is not None
    assert set(H.values()) <= set(written.headings)
    assert set(field_registry.IDENTITY_FIELDS) <= set(written.headings)


def test_an_export_read_back_into_the_same_database_creates_nothing(world, make_path):
    """The idempotency §17.1 is really asking for: a re-import is not a second plan."""
    make_path(PathStatus.DRAFT, code="RT-1")
    export = services.export_satnet_paths(
        actor=world["importer"], columns=[field.column for field in field_registry.INPUT_FIELDS]
    )
    batch = dry_run(world, [], content=export.content, batch_policy=BatchPolicy.ROW_BY_ROW)
    services.commit_batch(actor=world["importer"], batch=batch, content=export.content)

    assert SatnetPath.objects.count() == 1
    assert batch.rows.get().classification == RowClassification.DUPLICATE


# ---------------------------------------------------------------------------
# Authorization — administrator only (docs/design/03 §2.1)
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("role", [Role.OPERATOR, Role.APPROVER, Role.OBSERVER])
def test_only_an_administrator_may_run_a_dry_run(world, seeded_roles, role):
    user = make_user(f"imp-{role}", roles=[role])

    with pytest.raises(PermissionDenied):
        services.dry_run(actor=user, content=build_workbook([a_row(world)]), file_name="x.xlsx")


def test_only_an_administrator_may_commit(world, seeded_roles):
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content)
    operator = make_user("imp-op", roles=[Role.OPERATOR])

    with pytest.raises(PermissionDenied):
        services.commit_batch(actor=operator, batch=batch, content=content)

    assert not SatnetPath.objects.exists()


def test_a_denial_is_audited(world, seeded_roles):
    observer = make_user("imp-obs", roles=[Role.OBSERVER])

    with pytest.raises(PermissionDenied):
        services.dry_run(actor=observer, content=b"", file_name="x.xlsx")

    assert AuditEvent.objects.filter(action="PERMISSION_DENIED").exists()


def test_an_anonymous_caller_cannot_reach_the_import_screen(client):
    response = client.get(reverse("imports:list"))

    assert response.status_code in (302, 403)


def test_an_operator_is_refused_by_the_screen_as_well_as_by_the_service(client, seeded_roles):
    client.force_login(make_user("imp-op2", roles=[Role.OPERATOR]))

    assert client.get(reverse("imports:list")).status_code == 403


# ---------------------------------------------------------------------------
# The screens
# ---------------------------------------------------------------------------
def test_an_administrator_can_read_a_file_and_review_it(world, client):
    client.force_login(world["importer"])

    response = client.post(
        reverse("imports:dry-run"),
        {"file": _uploaded(build_workbook([a_row(world)])), "reason": "cutover rehearsal"},
        follow=True,
    )

    assert response.status_code == 200
    batch = ImportBatch.objects.get()
    assert batch.change_reason == "cutover rehearsal"
    assert b"IMP-1" in response.content


def test_the_review_screen_shows_every_classification_including_the_zeroes(world, client):
    """ "No conflicts" is a thing the reviewer needs to be told, not an absence to notice."""
    client.force_login(world["importer"])
    batch = dry_run(world, [a_row(world)])

    response = client.get(batch.get_absolute_url())

    assert response.status_code == 200
    for _value, label in RowClassification.choices:
        assert label.encode() in response.content


def test_committing_through_the_screen_needs_the_file_again(world, client):
    client.force_login(world["importer"])
    batch = dry_run(world, [a_row(world)])

    response = client.post(reverse("imports:commit", kwargs={"pk": batch.pk}), {}, follow=True)

    assert not SatnetPath.objects.exists()
    assert b"Attach the file again" in response.content


def test_committing_through_the_screen_writes_the_rows(world, client):
    client.force_login(world["importer"])
    content = build_workbook([a_row(world)])
    batch = dry_run(world, [], content=content)

    client.post(
        reverse("imports:commit", kwargs={"pk": batch.pk}),
        {"file": _uploaded(content)},
        follow=True,
    )

    assert SatnetPath.objects.get().code == "IMP-1"


def test_an_unreadable_upload_reports_its_reason_rather_than_raising(world, client):
    client.force_login(world["importer"])

    response = client.post(
        reverse("imports:dry-run"), {"file": _uploaded(b"not a workbook")}, follow=True
    )

    assert response.status_code == 200
    assert b"ZIP" in response.content
    assert not ImportBatch.objects.exists()


def test_the_navigation_offers_imports_to_an_administrator_only(world, client, seeded_roles):
    client.force_login(world["importer"])
    assert reverse("imports:list").encode() in client.get(reverse("home")).content

    client.force_login(make_user("imp-obs2", roles=[Role.OBSERVER]))
    assert reverse("imports:list").encode() not in client.get(reverse("home")).content


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _derived_heading(column_key: str) -> str:
    """The heading the export writes for a column the engine owns.

    Looked up rather than written out, for the same reason the input headings are: the importer
    reads what the exporter writes, and a test that restated the string would keep passing after
    the two had parted.
    """
    from reporting import columns as column_registry

    return normalized.heading(column_registry.BY_KEY[column_key])


def _uploaded(content: bytes):
    from django.core.files.uploadedfile import SimpleUploadedFile

    return SimpleUploadedFile(
        "plan.xlsx",
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
