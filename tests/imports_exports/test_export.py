"""The normalized export. §17.2, §21.12, §26.17, §26.19.

The first test in this file is the one the slice exists for. A Satnet Path code is data an
operator typed; Excel reads a cell beginning ``=`` as a *formula*, and formulas reach the shell
through DDE — so a code like ``=cmd|'/c calc'!A1`` is a working attack on whoever opens the
file. The platform is what has to stop it, and it stops it on the way *into* the cell rather
than hoping the reader's application is configured defensively.

Everything else here is about an export being trustworthy in the duller sense: it says who
produced it and from what, it carries the identifiers that let it come back through an import,
and it never shows anybody a row they could not already see.
"""

from __future__ import annotations

import dataclasses
import io
import uuid

import pytest
from django.core.exceptions import PermissionDenied
from django.urls import reverse
from openpyxl import load_workbook

from accounts.constants import Role
from audit.models import AuditEvent
from imports_exports import services
from imports_exports.export import legacy, normalized
from imports_exports.export.safety import is_dangerous, neutralise
from reporting import selectors as reporting_selectors
from satnet_paths.constants import PathStatus
from tests.factories import make_user

pytestmark = pytest.mark.django_db

MHZ = 1_000_000


@dataclasses.dataclass(frozen=True)
class Sheet:
    """One sheet, read eagerly."""

    headings: list
    rows: list[list]


def _read(content: bytes) -> dict[str, Sheet]:
    """Read a produced workbook into plain values, and close it.

    Eager rather than handing back a live workbook: openpyxl holds the underlying zip open, and
    a workbook collected later reports an unraisable exception that has nothing to do with the
    assertion that was being made.
    """
    workbook = load_workbook(io.BytesIO(content))
    try:
        return {
            name: Sheet(
                headings=[cell.value for cell in workbook[name][1]],
                rows=[list(row) for row in workbook[name].iter_rows(min_row=2, values_only=True)],
            )
            for name in workbook.sheetnames
        }
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# §21.12 — formula injection
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("prefix", ["=", "+", "-", "@", "\t", "\r"])
def test_a_value_that_would_be_evaluated_is_neutralised(prefix):
    dangerous = f"{prefix}cmd|'/c calc'!A1"

    assert is_dangerous(dangerous)
    assert not is_dangerous(neutralise(dangerous))


def test_an_ordinary_value_is_left_alone():
    """The guard must be narrow. An export where every cell grew an apostrophe would be
    unusable, and people would stop using the export rather than the apostrophe."""
    for value in ["SP-1", "29145000000", "Ka-band FWD", ""]:
        assert neutralise(value) == value


def test_a_negative_number_stays_a_number():
    """Numbers are written as typed cells and never reach the guard, so ``-5`` is not turned
    into text. A guard that mangled negatives would break every frequency offset in the file."""
    assert neutralise(-5) == -5
    assert neutralise(-5.5) == -5.5


def test_no_cell_in_a_produced_workbook_is_dangerous(lifecycle_world, make_path):
    """The property, asserted over the whole file rather than over the cells somebody
    remembered to check."""
    path = make_path(PathStatus.PLANNED, code="=cmd|'/c calc'!A1")

    content, _ = normalized.build(actor=lifecycle_world["admin"])

    for name, sheet in _read(content).items():
        for row in [sheet.headings, *sheet.rows]:
            for value in row:
                assert not is_dangerous(value), f"{name} still holds {value!r}"
    assert path.code.startswith("=")  # the record itself is untouched


def test_the_neutralised_value_is_still_recognisable(lifecycle_world, make_path):
    """Neutralising is not redacting. Whoever reads the file has to be able to tell which
    allocation the row is about."""
    make_path(PathStatus.PLANNED, code="=SUM(A1:A2)")

    content, _ = normalized.build(actor=lifecycle_world["admin"])

    codes = [row[2] for row in _read(content)["Satnet Paths"].rows]
    assert "'=SUM(A1:A2)" in codes


# ---------------------------------------------------------------------------
# What the file contains
# ---------------------------------------------------------------------------
def test_the_export_holds_one_row_per_current_allocation(lifecycle_world, make_path):
    make_path(PathStatus.PLANNED, code="E-1")
    make_path(PathStatus.DRAFT, code="E-2", centre=20 * MHZ)

    content, count = normalized.build(actor=lifecycle_world["admin"])

    assert count == 2
    assert len(_read(content)["Satnet Paths"].rows) == 2


def test_the_export_is_the_table(lifecycle_world, make_path):
    """Same filters, same scope-filtered queryset. An export answering a slightly different
    question from the screen is worse than none: somebody reconciles the difference by hand and
    concludes the platform is wrong."""
    make_path(PathStatus.PLANNED, code="E-LIVE")
    make_path(PathStatus.DRAFT, code="E-DRAFT", centre=20 * MHZ)
    admin = lifecycle_world["admin"]
    filters = {"status": PathStatus.PLANNED}

    _, count = normalized.build(actor=admin, filters=filters)

    assert count == reporting_selectors.table(admin, filters=filters).count()


def test_identifiers_are_always_exported_and_round_trip(lifecycle_world, make_path):
    """§17.2. Without the identifier an import can only ever create duplicates."""
    path = make_path(PathStatus.PLANNED, code="E-1")

    content, _ = normalized.build(actor=lifecycle_world["admin"], columns=["code"])

    sheet = _read(content)["Satnet Paths"]
    assert sheet.headings[:2] == ["id", "revision_group"]

    exported_id, exported_group = sheet.rows[0][:2]
    assert uuid.UUID(exported_id) == path.id
    assert uuid.UUID(exported_group) == path.revision_group


def test_a_frequency_is_exported_in_hz_rather_than_rendered(lifecycle_world, make_path):
    """**A-08**. A spreadsheet that received "29,145.000" would have to parse a thousands
    separator to get a number back."""
    path = make_path(PathStatus.PLANNED, code="E-1")

    content, _ = normalized.build(actor=lifecycle_world["admin"], columns=["occupied_bw"])

    value = _read(content)["Satnet Paths"].rows[0][2]
    assert value == path.occupied_bw_hz
    assert isinstance(value, int)


def test_chosen_columns_are_what_is_exported(lifecycle_world, make_path):
    make_path(PathStatus.PLANNED, code="E-1")

    content, _ = normalized.build(
        actor=lifecycle_world["admin"], columns=["code", "status", "occupied_bw"]
    )

    assert _read(content)["Satnet Paths"].headings == [
        "id",
        "revision_group",
        "Satnet Path",
        "Status",
        "OCCUPIED_BANDWIDTH",
    ]


def test_a_specification_column_is_headed_by_its_code(lifecycle_world, make_path):
    """§10.3 uses the code as the compact representation, and the Data Dictionary sheet
    explains it. A heading an administrator had renamed would stop matching the importer."""
    make_path(PathStatus.PLANNED, code="E-1")

    content, _ = normalized.build(actor=lifecycle_world["admin"], columns=["occupied_bw"])

    assert "OCCUPIED_BANDWIDTH" in _read(content)["Satnet Paths"].headings


# ---------------------------------------------------------------------------
# The Data Dictionary and Export sheets
# ---------------------------------------------------------------------------
def test_the_data_dictionary_comes_from_the_dictionary(lifecycle_world, make_path):
    """§2, §26.19. Not restated in the export, or it would go stale the first time an
    administrator improved a description."""
    from specifications.models import SpecificationDefinition

    make_path(PathStatus.PLANNED, code="E-1")
    definition = SpecificationDefinition.objects.get(code="OCCUPIED_BANDWIDTH")

    content, _ = normalized.build(actor=lifecycle_world["admin"], columns=["occupied_bw"])

    row = _read(content)["Data Dictionary"].rows[0]
    assert row[0] == "OCCUPIED_BANDWIDTH"
    assert row[1] == definition.display_name
    assert row[4] == definition.description


def test_an_edited_description_reaches_the_next_export(lifecycle_world, make_path):
    from specifications.models import SpecificationDefinition

    make_path(PathStatus.PLANNED, code="E-1")
    SpecificationDefinition.objects.filter(code="OCCUPIED_BANDWIDTH").update(
        description="Edited by an administrator."
    )

    content, _ = normalized.build(actor=lifecycle_world["admin"], columns=["occupied_bw"])

    assert _read(content)["Data Dictionary"].rows[0][4] == "Edited by an administrator."


def test_the_workbook_records_what_produced_it(lifecycle_world, make_path):
    """§17.2. An export without its filter parameters is a number nobody can reproduce."""
    make_path(PathStatus.PLANNED, code="E-1")

    content, _ = normalized.build(
        actor=lifecycle_world["admin"],
        filters={"status": PathStatus.PLANNED},
        columns=["code"],
    )

    recorded = dict(_read(content)["Export"].rows)
    assert recorded["Exported by"] == lifecycle_world["admin"].username
    assert recorded["Rows"] == 1
    assert recorded["Filter: status"] == PathStatus.PLANNED
    assert "UTC" in "".join(recorded)


def test_an_unfiltered_export_says_so(lifecycle_world, make_path):
    """The absence of filters is a fact worth recording: it is the difference between "these
    are all of them" and "these are the ones somebody narrowed to"."""
    make_path(PathStatus.PLANNED, code="E-1")

    content, _ = normalized.build(actor=lifecycle_world["admin"])

    recorded = dict(_read(content)["Export"].rows)
    assert "none" in recorded["Filter"]


def test_the_workbook_has_exactly_the_three_sheets(lifecycle_world, make_path):
    make_path(PathStatus.PLANNED, code="E-1")

    content, _ = normalized.build(actor=lifecycle_world["admin"])

    assert list(_read(content)) == ["Satnet Paths", "Data Dictionary", "Export"]


# ---------------------------------------------------------------------------
# Scope, capability and the trail
# ---------------------------------------------------------------------------
def test_an_export_never_widens_what_somebody_can_see(lifecycle_world, make_path):
    """`docs/design/03` §4: scope is applied at the queryset, not at render time."""
    from django.contrib.auth.models import AnonymousUser

    make_path(PathStatus.PLANNED, code="E-1")

    _, as_admin = normalized.build(actor=lifecycle_world["admin"])
    _, as_nobody = normalized.build(actor=AnonymousUser())

    assert as_admin == 1
    assert as_nobody == 0


def test_every_role_may_export(lifecycle_world, make_path):
    make_path(PathStatus.PLANNED, code="E-1")

    for role in (Role.ADMIN, Role.OPERATOR, Role.APPROVER, Role.OBSERVER):
        export = services.export_satnet_paths(actor=make_user(f"e-{role}", roles=[role]))
        assert export.row_count == 1


def test_an_anonymous_caller_is_refused(lifecycle_world):
    from django.contrib.auth.models import AnonymousUser

    with pytest.raises(PermissionDenied):
        services.export_satnet_paths(actor=AnonymousUser())


def test_every_export_is_audited(lifecycle_world, make_path):
    """§18. An export is the one action that puts operational data where the platform can no
    longer see it, so the trail records that it happened — the filters and the count, not the
    rows, because the trail is a record and not a second copy of the data."""
    make_path(PathStatus.PLANNED, code="E-1")

    services.export_satnet_paths(
        actor=lifecycle_world["admin"], filters={"status": PathStatus.PLANNED}, reason="Monthly"
    )

    event = AuditEvent.objects.filter(action="EXPORT_RUN").latest("occurred_at")
    assert event.after["rows"] == 1
    assert event.after["filters"] == {"status": PathStatus.PLANNED}
    assert event.change_reason == "Monthly"


# ---------------------------------------------------------------------------
# Over HTTP
# ---------------------------------------------------------------------------
def test_the_download_is_an_xlsx_attachment(client, lifecycle_world, make_path):
    make_path(PathStatus.PLANNED, code="E-1")
    client.force_login(lifecycle_world["admin"])

    response = client.get(reverse("exports:satnet-paths"))

    assert response.status_code == 200
    assert response["Content-Type"].endswith("spreadsheetml.sheet")
    assert "attachment; filename=" in response["Content-Disposition"]
    assert response["X-Export-Rows"] == "1"


def test_the_download_honours_the_table_filters(client, lifecycle_world, make_path):
    make_path(PathStatus.PLANNED, code="E-LIVE")
    make_path(PathStatus.DRAFT, code="E-DRAFT", centre=20 * MHZ)
    client.force_login(lifecycle_world["admin"])

    response = client.get(reverse("exports:satnet-paths"), {"status": PathStatus.DRAFT})

    assert response["X-Export-Rows"] == "1"
    codes = [row[2] for row in _read(response.content)["Satnet Paths"].rows]
    assert codes == ["E-DRAFT"]


def test_the_table_offers_the_export(client, lifecycle_world, make_path):
    make_path(PathStatus.PLANNED, code="E-1")
    client.force_login(lifecycle_world["admin"])

    body = client.get(reverse("reporting:satnet-paths"), {"status": PathStatus.PLANNED}).content

    assert b"/exports/satnet-paths.xlsx?status=PLANNED" in body


def test_an_anonymous_download_is_a_redirect_to_sign_in(client):
    response = client.get(reverse("exports:satnet-paths"))

    assert response.status_code == 302
    assert "/accounts/login/" in response["Location"]


# ---------------------------------------------------------------------------
# OQ-18 — the legacy export
# ---------------------------------------------------------------------------
def test_the_legacy_export_refuses_rather_than_guessing():
    """**OQ-18**. A layout written from a description looks approximately right and is wrong in
    ways nobody notices until a Phase 9 migration comparison blames the engine for it.

    A stub returning an empty workbook would be worse than this: somebody would ship it, and an
    empty file is indistinguishable from a filter that matched nothing.
    """
    with pytest.raises(legacy.LegacyExportUnavailable, match="sample of the incumbent workbook"):
        legacy.build()
